import asyncio
import websockets
import json
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
import base64
import io
import threading
import sqlite3
import csv
from datetime import datetime
import os
import csv
from datetime import datetime
import os

class ClientExporter:
    def __init__(self, db_connection=None):
        self.db_connection = db_connection
        self.export_dir = "exports"
        
        # Создаем директорию для экспортов, если её нет
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)

    def export_to_csv(self, clients_data):
        """
        Экспортирует данные клиентов в CSV файл
        
        Args:
            clients_data (dict): Словарь с данными клиентов
        
        Returns:
            str: Путь к созданному файлу
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(self.export_dir, f"clients_export_{timestamp}.csv")
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['UUID', 'Имя пользователя', 'Операционная система', 
                            'Процессор', 'ОЗУ (ГБ)', 'Разрешение экрана', 'Филиал']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                for uuid, client in clients_data.items():
                    writer.writerow({
                        'UUID': uuid,
                        'Имя пользователя': client.get('username', 'Неизвестно'),
                        'Операционная система': client.get('os', 'Неизвестно'),
                        'Процессор': client.get('processor', 'Неизвестно'),
                        'ОЗУ (ГБ)': client.get('ram_total', 'Неизвестно'),
                        'Разрешение экрана': f"{client.get('screen_resolution', {}).get('width', '?')}x{client.get('screen_resolution', {}).get('height', '?')}",
                        'Филиал': client.get('branch', 'Не указан')
                    })
                
                return filename
        except Exception as e:
            print(f"Ошибка при экспорте: {e}")
            return None

    def export_to_excel(self, clients_data):
        """
        Экспортирует данные клиентов в Excel файл
        (требует установки библиотеки openpyxl)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Клиенты"
            
            # Заголовки
            headers = ['UUID', 'Имя пользователя', 'Операционная система', 
                      'Процессор', 'ОЗУ (ГБ)', 'Разрешение экрана', 'Филиал']
            
            # Стили
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Записываем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Записываем данные
            row = 2
            for uuid, client in clients_data.items():
                ws.cell(row=row, column=1, value=uuid)
                ws.cell(row=row, column=2, value=client.get('username', 'Неизвестно'))
                ws.cell(row=row, column=3, value=client.get('os', 'Неизвестно'))
                ws.cell(row=row, column=4, value=client.get('processor', 'Неизвестно'))
                ws.cell(row=row, column=5, value=client.get('ram_total', 'Неизвестно'))
                ws.cell(row=row, column=6, value=f"{client.get('screen_resolution', {}).get('width', '?')}x{client.get('screen_resolution', {}).get('height', '?')}")
                ws.cell(row=row, column=7, value=client.get('branch', 'Не указан'))
                row += 1
            
            # Автоматическая ширина столбцов
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(self.export_dir, f"clients_export_{timestamp}.xlsx")
            wb.save(filename)
            return filename
            
        except ImportError:
            print("Для экспорта в Excel требуется установить библиотеку openpyxl")
            return None
        except Exception as e:
            print(f"Ошибка при экспорте в Excel: {e}")
            return None
# Глобальные переменные
clients = {}  # Словарь для хранения подключённых клиентов
selected_client_id = None  # Выбранный клиент
loop = None  # Глобальный event loop для асинхронных операций
update_task = None  # Задача для периодического обновления скриншотов
speed_label = None
info_label = None
is_control_enabled = False
def init_database():
    conn = sqlite3.connect("clients.db")
    cursor = conn.cursor()
    
    # Создание таблицы для хранения информации о клиентах
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            uuid TEXT UNIQUE,  -- Уникальный идентификатор клиента
            username TEXT,
            os TEXT,
            processor TEXT,
            ram_total REAL,
            screen_width INTEGER,
            screen_height INTEGER,
            branch TEXT  -- Название филиала
        )
    """)
    
    conn.commit()
    conn.close()
def save_client_to_db(client_info, branch=None):
    try:
        conn = sqlite3.connect("clients.db")
        cursor = conn.cursor()
        
        client_uuid = client_info.get("uuid")
        if not client_uuid:
            print("Ошибка: отсутствует UUID клиента")
            return False

        # Если branch не передан, используем значение из client_info
        if branch is None:
            branch = client_info.get("branch")

        # Проверяем существование клиента
        cursor.execute("SELECT branch FROM clients WHERE uuid = ?", (client_uuid,))
        existing_client = cursor.fetchone()
        
        if existing_client:
            # Обновляем существующего клиента
            cursor.execute("""
                UPDATE clients 
                SET username = ?, 
                    os = ?, 
                    processor = ?, 
                    ram_total = ?, 
                    screen_width = ?, 
                    screen_height = ?, 
                    branch = ?
                WHERE uuid = ?
            """, (
                client_info["username"],
                client_info["os"],
                client_info["processor"],
                client_info["ram_total"],
                client_info["screen_resolution"]["width"],
                client_info["screen_resolution"]["height"],
                branch,  # Используем переданный или полученный branch
                client_uuid
            ))
        else:
            # Добавляем нового клиента
            cursor.execute("""
                INSERT INTO clients 
                (uuid, username, os, processor, ram_total, screen_width, screen_height, branch)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                client_uuid,
                client_info["username"],
                client_info["os"],
                client_info["processor"],
                client_info["ram_total"],
                client_info["screen_resolution"]["width"],
                client_info["screen_resolution"]["height"],
                branch
            ))
        
        conn.commit()
        print(f"Клиент {client_info['username']} сохранен с филиалом {branch}")
        return True
        
    except sqlite3.Error as e:
        print(f"Ошибка базы данных: {e}")
        return False
    finally:
        conn.close()

def load_clients_from_db():
    try:
        conn = sqlite3.connect("clients.db")
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT uuid, username, os, processor, ram_total, 
                   screen_width, screen_height, branch 
            FROM clients
        """)
        
        clients_data = {}
        for row in cursor.fetchall():
            uuid, username, os, processor, ram_total, sw, sh, branch = row
            clients_data[uuid] = {
                "username": username,
                "os": os,
                "processor": processor,
                "ram_total": ram_total,
                "screen_resolution": {
                    "width": sw,
                    "height": sh
                },
                "branch": branch
            }
        
        return clients_data
        
    except sqlite3.Error as e:
        print(f"Ошибка при загрузке клиентов: {e}")
        return {}
    finally:
        conn.close()
init_database()
saved_clients = load_clients_from_db()

for username, info in saved_clients.items():
    print(f"Загружен клиент из базы данных: {username}, {info}")
# Отправка команды клиенту
async def send_command(client_id, command):
    if client_id in clients and "websocket" in clients[client_id]:
        websocket = clients[client_id]["websocket"]
        await websocket.send(command)
        print(f"Команда отправлена клиенту {client_id}: {command}")
    else:
        print(f"Клиент {client_id} не найден или соединение отсутствует.")

# Периодическая отправка скриншотов
async def send_periodic_screenshots(client_id, interval=3):  # Интервал в секундах
    while True:
        await send_command(client_id, "screenshot")
        await asyncio.sleep(interval)

# Обработчик подключений
async def handle_client(websocket):
    global selected_client_id
    client_id = len(clients) + 1
    clients[client_id] = {"websocket": websocket, "info": {}}
    
    try:
        async for message in websocket:
            if message.startswith("system_info:"):
                _, info_str = message.split(":", 1)
                system_info = json.loads(info_str)
                
                # Проверяем, есть ли информация о клиенте в базе данных
                current_clients = load_clients_from_db()
                client_uuid = system_info.get("uuid")
                
                if client_uuid and client_uuid in current_clients:
                    # Если клиент существует, сохраняем его branch
                    existing_branch = current_clients[client_uuid].get("branch")
                    if existing_branch:
                        system_info["branch"] = existing_branch
                        print(f"Загружен существующий филиал: {existing_branch} для клиента {system_info['username']}")
                
                # Сохраняем информацию в словарь клиентов
                clients[client_id]["info"] = system_info
                
                # Сохраняем/обновляем информацию в базе данных
                save_client_to_db(system_info)
                
                # Обновляем GUI
                root.after(0, update_client_list)
                
                if info_label:
                    branch_info = f"Филиал: {system_info.get('branch', 'Не указан')}"
                    root.after(0, lambda: info_label.config(
                        text=f"Имя: {system_info['username']}, "
                             f"OS: {system_info['os']}, "
                             f"Процессор: {system_info['processor']}, "
                             f"RAM: {system_info['ram_total']} ГБ, "
                             f"Экран: {system_info['screen_resolution']['width']}x"
                             f"{system_info['screen_resolution']['height']}\n"
                             f"{branch_info}"
                    ))
                
            # ... остальной код обработки сообщений ...
                
                # Обновляем список клиентов в реальном времени
                update_client_list()
            
            elif message == "ping":
                await websocket.send(str(asyncio.get_event_loop().time()))
            
            elif message.startswith("pong:"):
                _, timestamp = message.split(":")
                server_time = asyncio.get_event_loop().time()
                client_time = float(timestamp)
                latency = (server_time - client_time) * 1000
                print(f"Пинг клиента {client_id}: {latency:.2f} мс")
                if speed_label:
                    speed_label.config(text=f"Пинг: {latency:.2f} мс")
            
            else:
                # Если это не команда, а данные скриншота
                if selected_client_id == client_id:
                    update_screenshot(message)
    
    except websockets.exceptions.ConnectionClosed:
        print(f"Клиент {client_id} отключился.")
        del clients[client_id]
        root.after(0, update_client_list)  # Безопасное обновление GUI # Обновляем список клиентов после отключения
# Обновление скриншота в GUI
def update_screenshot(data):
    try:
        print(f"Получен скриншот. Размер данных: {len(data)} байт.")
        screenshot_data = base64.b64decode(data)
        image = Image.open(io.BytesIO(screenshot_data))
        photo = ImageTk.PhotoImage(image.resize((800, 600), Image.LANCZOS))
        screenshot_label.config(image=photo)
        screenshot_label.image = photo  # Сохраняем ссылку на изображение
    except Exception as e:
        print(f"Ошибка при обновлении скриншота: {e}")

# Обновление списка клиентов
def update_client_list():
    client_list.delete(0, tk.END)  # Очищаем текущий список
    
    # Получаем актуальные данные из базы
    current_clients = load_clients_from_db()
    
    # Создаем список клиентов с их данными для сортировки
    client_entries = []
    for uuid, client_data in current_clients.items():
        username = client_data.get("username", "Неизвестный")
        os_info = client_data.get("os", "Неизвестная ОС")
        branch = client_data.get("branch", "999")  # Используем '999' для клиентов без филиала
        
        # Проверяем, активен ли клиент
        is_active = any(
            client.get("info", {}).get("uuid") == uuid 
            for client in clients.values()
        )
        
        status = "✔️ " if is_active else "❌ "
        
        # Создаем кортеж с данными для сортировки
        client_entries.append((
            branch,  # Первый элемент для сортировки
            {
                'status': status,
                'username': username,
                'os_info': os_info,
                'branch': branch
            }
        ))
    
    # Сортируем клиентов по номеру филиала
    client_entries.sort(key=lambda x: (
        '999' if x[0] is None else x[0],  # Сначала сортируем по филиалу
        x[1]['username']  # Затем по имени пользователя
    ))
    
    # Добавляем отсортированных клиентов в список
    for _, client in client_entries:
        branch_display = f"[Филиал {client['branch']}]" if client['branch'] != '999' else '[Нет филиала]'
        display_text = f"{client['status']}{branch_display} {client['username']} | {client['os_info']}"
        client_list.insert(tk.END, display_text)

    print("Список клиентов обновлен")
# Выбор клиента из списка
def select_client(event):
    global selected_client_id, update_task
    selection = client_list.curselection()
    if not selection:
        print("Нет выбранного клиента")
        return
        
    # Получаем актуальные данные о клиентах
    current_clients = load_clients_from_db()
    client_uuids = list(current_clients.keys())
    
    try:
        selected_index = selection[0]
        selected_uuid = client_uuids[selected_index]
        
        # Находим ID активного клиента по UUID
        for client_id, client_data in clients.items():
            if client_data.get("info", {}).get("uuid") == selected_uuid:
                selected_client_id = client_id
                print(f"Выбран клиент {client_id} (UUID: {selected_uuid})")
                
                # Запускаем периодическое обновление скриншотов
                try:
                    interval = float(update_interval_entry.get())
                    if interval <= 0:
                        raise ValueError("Интервал должен быть больше 0")
                except ValueError:
                    print("Ошибка: Используется значение по умолчанию (3 секунды)")
                    interval = 3
                
                if update_task is not None:
                    update_task.cancel()
                update_task = asyncio.run_coroutine_threadsafe(
                    send_periodic_screenshots(selected_client_id, interval), 
                    loop
                )
                return
                
        print("Выбранный клиент не активен")
        selected_client_id = None
        
    except IndexError:
        print("Ошибка: неверный индекс клиента")
        selected_client_id = None
# Перемещение курсора
def move_mouse(x, y):
    if selected_client_id is not None:
        asyncio.run_coroutine_threadsafe(
            send_command(selected_client_id, f"move_mouse {x} {y}"), loop
        )

# Нажатие клавиши
def press_key(key):
    if selected_client_id is not None:
        asyncio.run_coroutine_threadsafe(
            send_command(selected_client_id, f"press_key {key}"), loop
        )

# Создание GUI
def create_gui():
    
    global root,client_list, screenshot_label, update_interval_entry, speed_label, info_label, control_button, branch_entry

    root = tk.Tk()
    root.title("Удалённое управление")
    root.geometry("1200x800")  # Устанавливаем размер окна
    root.configure(bg="#1e1e1e")  # Темная тема

    # Настройка стиля
    style = ttk.Style()
    style.theme_use("clam")  # Используем стиль "clam" для более современного вида
    style.configure("TFrame", background="#1e1e1e")
    style.configure("TLabel", background="#1e1e1e", foreground="#ffffff", font=("Roboto", 12))
    style.configure("TButton", background="#007acc", foreground="#ffffff", font=("Roboto", 10), padding=5)
    style.map("TButton", background=[("active", "#005f99")])
    style.configure("TEntry", fieldbackground="#333333", foreground="#ffffff", font=("Roboto", 10))

    # Список клиентов
    client_frame = ttk.Frame(root)
    client_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

    client_list = tk.Listbox(client_frame, width=25, height=20, bg="#333333", fg="#ffffff", font=("Roboto", 10))
    client_list.pack(fill=tk.BOTH, expand=True)
    client_list.bind("<<ListboxSelect>>", select_client)

    # Просмотр скриншота
    screenshot_frame = ttk.Frame(root)
    screenshot_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

    screenshot_label = tk.Label(screenshot_frame, bg="#2d2d2d", relief="solid", bd=1)
    screenshot_label.pack(fill=tk.BOTH, expand=True)

    # Метка для отображения информации о клиенте
    info_frame = ttk.Frame(root)
    info_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)

    info_label = ttk.Label(info_frame, text="Информация о клиенте: Нет данных", font=("Roboto", 10))
    info_label.pack(side=tk.LEFT)

    # Поле для ввода филиала
    branch_frame = ttk.Frame(root)
    branch_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)

    branch_label = ttk.Label(branch_frame, text="Филиал:", font=("Roboto", 10))
    branch_label.pack(side=tk.LEFT)

    branch_entry = ttk.Entry(branch_frame, width=20, font=("Roboto", 10))
    branch_entry.pack(side=tk.LEFT, padx=5)

    save_branch_button = ttk.Button(branch_frame, text="Сохранить филиал",
                                    command=lambda: save_branch_for_client())
    save_branch_button.pack(side=tk.LEFT, padx=5)
    def periodic_update():
        update_client_list()
        root.after(5000, periodic_update)  # Обновление каждые 5 секунд
    
    # Запускаем периодическое обновление
    periodic_update()
    # Функция сохранения филиала
    # Функция сохранения филиала
    def save_branch_for_client():
        selection = client_list.curselection()
        if not selection:
            print("Ошибка: Не выбран клиент в списке")
            return
        
        # Получаем название филиала
        branch_name = branch_entry.get().strip()
        if not branch_name:
            print("Ошибка: Введите название филиала")
            return

        # Получаем UUID выбранного активного клиента
        selected_index = selection[0]
        
        # Сначала проверяем активных клиентов
        for client_id, client_data in clients.items():
            client_info = client_data.get("info", {})
            if client_info:  # Проверяем, что у клиента есть информация
                print(f"Проверка клиента {client_id}: {client_info}")
                
                # Получаем UUID активного клиента
                client_uuid = client_info.get("uuid")
                if client_uuid:
                    print(f"Найден UUID: {client_uuid}")
                    
                    # Обновляем информацию о филиале
                    client_info['branch'] = branch_name
                    
                    # Сохраняем в базу данных
                    if save_client_to_db(client_info, branch=branch_name):
                        print(f"Филиал '{branch_name}' сохранен для клиента {client_info.get('username')}")
                        update_client_list()
                        update_branch_info(None)
                        return
                    else:
                        print("Ошибка при сохранении филиала")
                        return

        print("Не удалось найти активного клиента с UUID")
    branch_status_frame = ttk.Frame(root)
    branch_status_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)

    branch_info_label = ttk.Label(branch_status_frame, text="Филиал: Не указан", 
                                 font=("Roboto", 10))
    branch_info_label.pack(anchor=tk.CENTER)
    def update_branch_info(event=None):
        selection = client_list.curselection()
        if selection:
            current_clients = load_clients_from_db()
            client_uuids = list(current_clients.keys())
            try:
                selected_uuid = client_uuids[selection[0]]
                client_data = current_clients[selected_uuid]
                branch = client_data.get('branch', 'Не указан')
                branch_info_label.config(
                    text=f"Филиал: {branch}",
                    foreground='#2ecc71' if branch != 'Не указан' else '#e74c3c'
                )
            except (IndexError, KeyError):
                branch_info_label.config(
                    text="Филиал: Ошибка", 
                    foreground='#e74c3c'
                )
        else:
            branch_info_label.config(
                text="Филиал: Не выбран", 
                foreground='#95a5a6'
            )
    client_list.bind("<<ListboxSelect>>", lambda e: (select_client(e), update_branch_info(e)))
    # Настройка интервала обновления
    
    # Создаем фрейм для кнопок экспорта
    export_frame = ttk.Frame(root)
    export_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=5)

    def export_clients():
        current_clients = load_clients_from_db()
        if not current_clients:
            print("Нет данных для экспорта")
            return
            
        exporter = ClientExporter()
        
        # Экспорт в CSV
        csv_file = exporter.export_to_csv(current_clients)
        if csv_file:
            print(f"Данные экспортированы в CSV: {csv_file}")
        
        # Экспорт в Excel (если установлен openpyxl)
        excel_file = exporter.export_to_excel(current_clients)
        if excel_file:
            print(f"Данные экспортированы в Excel: {excel_file}")

    export_button = ttk.Button(export_frame, text="Экспорт клиентов",
                              command=export_clients)
    export_button.pack(side=tk.LEFT, padx=5)
    interval_frame = ttk.Frame(root)
    interval_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)

    interval_label = ttk.Label(interval_frame, text="Интервал обновления (сек):", font=("Roboto", 10))
    interval_label.pack(side=tk.LEFT)

    update_interval_entry = ttk.Entry(interval_frame, width=5, font=("Roboto", 10))
    update_interval_entry.insert(0, "3")  # Значение по умолчанию
    update_interval_entry.pack(side=tk.LEFT, padx=5)

    # Метка для отображения пинга
    speed_frame = ttk.Frame(root)
    speed_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)

    speed_label = ttk.Label(speed_frame, text="Пинг: Нет данных", font=("Roboto", 10))
    speed_label.pack(side=tk.LEFT)

    # Контроллеры для управления
    control_frame = ttk.Frame(root)
    control_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)

    move_mouse_button = ttk.Button(control_frame, text="Переместить курсор",
                                   command=lambda: move_mouse(100, 100))
    move_mouse_button.pack(side=tk.LEFT, padx=5)

    press_key_button = ttk.Button(control_frame, text="Нажать клавишу",
                                  command=lambda: press_key("enter"))
    press_key_button.pack(side=tk.LEFT, padx=5)

    ping_button = ttk.Button(control_frame, text="Проверить пинг",
                             command=lambda: asyncio.run_coroutine_threadsafe(send_command(selected_client_id, "ping"), loop))
    ping_button.pack(side=tk.LEFT, padx=5)
  
    # Кнопка-тумблер для управления экраном
    def toggle_control():
        global is_control_enabled
        is_control_enabled = not is_control_enabled
        if is_control_enabled:
            control_button.config(text="Управление ВКЛ", style="Accent.TButton")
        else:
            control_button.config(text="Управление ВЫКЛ", style="TButton")

    style.configure("Accent.TButton", background="#2ecc71", foreground="#ffffff", font=("Roboto", 10))
    style.map("Accent.TButton", background=[("active", "#27ae60")])

    control_button = ttk.Button(control_frame, text="Управление ВЫКЛ", command=toggle_control)
    control_button.pack(side=tk.LEFT, padx=5)

    # Обработчики событий мыши
    def on_mouse_click(event):
        if selected_client_id is not None and is_control_enabled:
            # Получаем координаты клика относительно скриншота
            x, y = event.x, event.y
            print(f"Клик на скриншоте: ({x}, {y})")
            
            # Получаем разрешение экрана клиента
            client_data = clients.get(selected_client_id, {})
            screen_info = client_data.get("info", {}).get("screen_resolution", {"width": 1920, "height": 1080})
            screen_width = screen_info["width"]
            screen_height = screen_info["height"]

            # Пересчитываем координаты в реальные координаты экрана клиента
            label_width = screenshot_label.winfo_width()
            label_height = screenshot_label.winfo_height()
            real_x = int(x / label_width * screen_width)
            real_y = int(y / label_height * screen_height)
            
            # Отправляем команду на перемещение мыши
            asyncio.run_coroutine_threadsafe(
                send_command(selected_client_id, f"move_mouse {real_x} {real_y}"), loop
            )

    def on_mouse_press(event):
        if selected_client_id is not None and is_control_enabled:
            # Левый клик
            if event.num == 1:
                asyncio.run_coroutine_threadsafe(send_command(selected_client_id, "press_mouse left"), loop)
            # Правый клик
            elif event.num == 3:
                asyncio.run_coroutine_threadsafe(send_command(selected_client_id, "press_mouse right"), loop)
    root.after(100, update_client_list) 
    # Привязываем обработчики событий к скриншоту
    screenshot_label.bind("<Motion>", on_mouse_click)  # Движение мыши
    screenshot_label.bind("<Button-1>", on_mouse_press)  # Левый клик
    screenshot_label.bind("<Button-3>", on_mouse_press)  # Правый клик
    screenshot_label.bind("<ButtonRelease-1>", lambda _: None)  # Отпускание левой кноп

    # Запуск сервера в отдельном потоке
    server_thread = threading.Thread(target=start_server, daemon=True)
    server_thread.start()

    # Обновление списка клиентов
    update_client_list()

    root.mainloop()

# Запуск WebSocket-сервера

 
def start_server():
    global loop
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    async def serve():
        async with websockets.serve(handle_client, "0.0.0.0", 8765):
            print("Server started on ws://0.0.0.0:8765")
            await asyncio.Future()  # Keep the server running indefinitely
    loop.run_until_complete(serve())

if __name__ == "__main__":
    create_gui()